from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth, TruncDay
from django.utils import timezone
from django.conf import settings
from datetime import timedelta
import csv
import io
import logging
import stripe

from .models import SellerSubscription, SubscriptionBillingHistory, BulkImport, BulkImportRow, InventoryItem
from marketplace.models import Listing, Order, PaymentMethod
from items.models import Category

logger = logging.getLogger('seller_tools')

stripe.api_key = settings.STRIPE_SECRET_KEY


@login_required
def seller_dashboard(request):
    """Main seller dashboard"""
    # Get or create subscription
    subscription, created = SellerSubscription.objects.get_or_create(
        user=request.user,
        defaults={'tier': 'starter'}
    )

    # Stats
    active_listings = Listing.objects.filter(seller=request.user, status='active').count()
    pending_orders = Order.objects.filter(seller=request.user, status='paid').count()

    # Recent sales
    recent_sales = Order.objects.filter(
        seller=request.user,
        status__in=['paid', 'shipped', 'delivered', 'completed']
    ).order_by('-created')[:10]

    # Sales this month
    month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    monthly_sales = Order.objects.filter(
        seller=request.user,
        status__in=['paid', 'shipped', 'delivered', 'completed'],
        created__gte=month_start
    ).aggregate(
        total=Sum('item_price'),
        count=Count('id')
    )

    # Auction events accepting submissions (for trusted sellers)
    auction_events_accepting = []
    if hasattr(request.user, 'profile') and request.user.profile.is_trusted_seller:
        from marketplace.models import AuctionEvent
        auction_events_accepting = AuctionEvent.objects.filter(
            is_platform_event=True,
            accepting_submissions=True,
        ).order_by('submission_deadline')[:5]

    return render(request, 'seller_tools/dashboard.html', {
        'subscription': subscription,
        'active_listings': active_listings,
        'pending_orders': pending_orders,
        'recent_sales': recent_sales,
        'monthly_sales': monthly_sales,
        'auction_events_accepting': auction_events_accepting,
    })


@login_required
def subscription_manage(request):
    """Manage seller subscription"""
    subscription, created = SellerSubscription.objects.get_or_create(
        user=request.user,
        defaults={'tier': 'starter'}
    )

    tiers = SellerSubscription.TIER_DETAILS
    payment_methods = PaymentMethod.objects.filter(user=request.user)

    # Get recent billing history
    billing_history = SubscriptionBillingHistory.objects.filter(
        subscription=subscription
    ).order_by('-created')[:5]

    return render(request, 'seller_tools/subscription.html', {
        'subscription': subscription,
        'tiers': tiers,
        'payment_methods': payment_methods,
        'billing_history': billing_history,
        'stripe_public_key': settings.STRIPE_PUBLIC_KEY,
    })


@login_required
def subscription_upgrade(request, tier):
    """Upgrade seller subscription with direct payment"""
    from marketplace.services.subscription_service import SubscriptionService

    if tier not in ['basic', 'featured', 'premium']:
        messages.error(request, 'Invalid subscription tier')
        return redirect('seller_tools:subscription')

    subscription, _ = SellerSubscription.objects.get_or_create(
        user=request.user,
        defaults={'tier': 'starter'}
    )
    tier_info = SellerSubscription.TIER_DETAILS[tier]

    # Calculate proration if upgrading from another paid tier
    proration = SubscriptionService.calculate_proration(subscription, tier)

    if request.method == 'POST':
        payment_method_id = request.POST.get('payment_method_id')
        if not payment_method_id:
            messages.error(request, 'Please provide a payment method')
            return redirect('seller_tools:subscription_upgrade', tier=tier)

        try:
            SubscriptionService.subscribe(request.user, tier, payment_method_id)
            messages.success(request, f'Successfully subscribed to {tier_info["name"]}!')
            return redirect('seller_tools:subscription')
        except stripe.error.CardError as e:
            messages.error(request, f'Payment failed: {e.user_message}')
        except Exception as e:
            messages.error(request, f'Unable to process subscription: {e}')
            return redirect('seller_tools:subscription')

    # GET - show payment form
    payment_methods = PaymentMethod.objects.filter(user=request.user)

    return render(request, 'seller_tools/subscription_upgrade.html', {
        'subscription': subscription,
        'tier': tier,
        'tier_info': tier_info,
        'proration': proration,
        'payment_methods': payment_methods,
        'stripe_public_key': settings.STRIPE_PUBLIC_KEY,
    })


@login_required
def subscription_success(request):
    """Handle successful subscription signup"""
    messages.success(request, 'Welcome! Your subscription is now active.')
    return redirect('seller_tools:subscription')


@login_required
@require_POST
def subscription_cancel(request):
    """Cancel seller subscription"""
    from marketplace.services.subscription_service import SubscriptionService

    subscription = get_object_or_404(SellerSubscription, user=request.user)

    if subscription.tier == 'starter':
        messages.info(request, 'You are already on the Starter tier.')
        return redirect('seller_tools:subscription')

    try:
        # Cancel at period end (graceful cancellation)
        SubscriptionService.cancel(request.user, at_period_end=True)
        messages.success(request, 'Your subscription will be cancelled at the end of the billing period.')
    except Exception as e:
        messages.error(request, f'Unable to cancel subscription: {e}')

    return redirect('seller_tools:subscription')


@login_required
@require_POST
def subscription_reactivate(request):
    """Reactivate a subscription that was set to cancel"""
    from marketplace.services.subscription_service import SubscriptionService

    try:
        SubscriptionService.reactivate(request.user)
        messages.success(request, 'Your subscription has been reactivated.')
    except Exception as e:
        messages.error(request, f'Unable to reactivate subscription: {e}')

    return redirect('seller_tools:subscription')


@login_required
def subscription_payment_methods(request):
    """Manage payment methods for subscription"""
    from marketplace.services.subscription_service import SubscriptionService

    subscription, _ = SellerSubscription.objects.get_or_create(
        user=request.user,
        defaults={'tier': 'starter'}
    )
    payment_methods = PaymentMethod.objects.filter(user=request.user)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            payment_method_id = request.POST.get('payment_method_id')
            if payment_method_id:
                try:
                    # Get or create Stripe customer
                    customer = SubscriptionService.get_or_create_stripe_customer(request.user)

                    # Attach payment method
                    stripe.PaymentMethod.attach(payment_method_id, customer=customer.id)

                    # Get payment method details
                    pm = stripe.PaymentMethod.retrieve(payment_method_id)

                    # Save locally
                    new_pm = PaymentMethod.objects.create(
                        user=request.user,
                        stripe_payment_method_id=payment_method_id,
                        card_brand=pm.card.brand,
                        card_last4=pm.card.last4,
                        card_exp_month=pm.card.exp_month,
                        card_exp_year=pm.card.exp_year,
                        is_default=not payment_methods.exists(),
                    )

                    # Set as default for subscription if it's the only one
                    if new_pm.is_default and subscription.tier != 'starter':
                        subscription.default_payment_method = new_pm
                        subscription.save(update_fields=['default_payment_method'])

                    messages.success(request, 'Payment method added successfully.')
                except stripe.error.CardError as e:
                    messages.error(request, f'Card error: {e.user_message}')
                except Exception as e:
                    messages.error(request, f'Error adding payment method: {e}')

        elif action == 'set_default':
            pm_id = request.POST.get('payment_method_id')
            try:
                pm = PaymentMethod.objects.get(id=pm_id, user=request.user)
                PaymentMethod.objects.filter(user=request.user).update(is_default=False)
                pm.is_default = True
                pm.save(update_fields=['is_default'])

                # Update subscription default
                subscription.default_payment_method = pm
                subscription.save(update_fields=['default_payment_method'])

                messages.success(request, 'Default payment method updated.')
            except PaymentMethod.DoesNotExist:
                messages.error(request, 'Payment method not found.')

        elif action == 'delete':
            pm_id = request.POST.get('payment_method_id')
            try:
                pm = PaymentMethod.objects.get(id=pm_id, user=request.user)

                # Don't delete if it's the subscription's payment method
                if subscription.default_payment_method_id == pm.id:
                    messages.error(request, 'Cannot delete the payment method used for your subscription. Set a different default first.')
                else:
                    # Detach from Stripe
                    try:
                        stripe.PaymentMethod.detach(pm.stripe_payment_method_id)
                    except Exception:
                        pass
                    pm.delete()
                    messages.success(request, 'Payment method removed.')
            except PaymentMethod.DoesNotExist:
                messages.error(request, 'Payment method not found.')

        return redirect('seller_tools:subscription_payment_methods')

    return render(request, 'seller_tools/subscription_payment_methods.html', {
        'subscription': subscription,
        'payment_methods': payment_methods,
        'stripe_public_key': settings.STRIPE_PUBLIC_KEY,
    })


@login_required
def subscription_billing_history(request):
    """View subscription billing history"""
    subscription = get_object_or_404(SellerSubscription, user=request.user)

    billing_history = SubscriptionBillingHistory.objects.filter(
        subscription=subscription
    ).order_by('-created')

    return render(request, 'seller_tools/subscription_billing_history.html', {
        'subscription': subscription,
        'billing_history': billing_history,
    })


@login_required
def bulk_import_list(request):
    """List all bulk imports"""
    imports = BulkImport.objects.filter(user=request.user)

    return render(request, 'seller_tools/import_list.html', {
        'imports': imports,
    })


@login_required
def bulk_import_create(request):
    """Create a new bulk import"""
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'Please upload a file')
            return redirect('seller_tools:import_create')

        file = request.FILES['file']
        file_type = file.name.split('.')[-1].lower()

        if file_type not in ['csv', 'xlsx']:
            messages.error(request, 'Only CSV and Excel files are supported')
            return redirect('seller_tools:import_create')

        bulk_import = BulkImport.objects.create(
            user=request.user,
            file=file,
            file_name=file.name,
            file_type=file_type,
            auto_publish=request.POST.get('auto_publish') == 'on',
        )

        # Save any uploaded image files for matching by filename
        from django.core.files.storage import default_storage
        images = request.FILES.getlist('images')
        if images:
            for img_file in images:
                path = f'bulk_imports/{bulk_import.id}/images/{img_file.name}'
                default_storage.save(path, img_file)

        # Parse the file and count rows
        if file_type == 'csv':
            content = file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(content)
            rows = list(reader)
            bulk_import.total_rows = len(rows)

            for i, row in enumerate(rows, 1):
                BulkImportRow.objects.create(
                    bulk_import=bulk_import,
                    row_number=i,
                    data=dict(row)
                )

            bulk_import.status = 'validating'
            bulk_import.save()

        elif file_type == 'xlsx':
            import openpyxl

            file.seek(0)
            try:
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            except Exception as e:
                bulk_import.delete()
                messages.error(request, f'Could not read Excel file: {e}')
                return redirect('seller_tools:import_create')

            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)

            # First row is headers
            try:
                header_row = next(rows_iter)
            except StopIteration:
                bulk_import.delete()
                messages.error(request, 'Excel file is empty')
                return redirect('seller_tools:import_create')

            headers = [str(h).strip().lower() if h else '' for h in header_row]
            row_count = 0

            for i, row_values in enumerate(rows_iter, 1):
                # Skip fully empty rows
                if not any(v is not None and str(v).strip() for v in row_values):
                    continue

                row_data = {}
                for col_idx, header in enumerate(headers):
                    if not header:
                        continue
                    val = row_values[col_idx] if col_idx < len(row_values) else None
                    row_data[header] = str(val).strip() if val is not None else ''

                BulkImportRow.objects.create(
                    bulk_import=bulk_import,
                    row_number=i,
                    data=row_data,
                )
                row_count += 1

            wb.close()
            bulk_import.total_rows = row_count
            bulk_import.status = 'validating'
            bulk_import.save()

        messages.success(request, f'Imported {bulk_import.total_rows} rows. Review and process.')
        return redirect('seller_tools:import_detail', pk=bulk_import.pk)

    categories = Category.objects.filter(parent__isnull=True)

    return render(request, 'seller_tools/import_create.html', {
        'categories': categories,
    })


@login_required
def bulk_import_detail(request, pk):
    """View bulk import details"""
    bulk_import = get_object_or_404(BulkImport, pk=pk, user=request.user)
    rows = bulk_import.rows.all()[:100]  # Limit for display

    return render(request, 'seller_tools/import_detail.html', {
        'bulk_import': bulk_import,
        'rows': rows,
    })


@login_required
@require_POST
def bulk_import_process(request, pk):
    """Process bulk import and create listings"""
    bulk_import = get_object_or_404(BulkImport, pk=pk, user=request.user)

    if bulk_import.status not in ['validating', 'partial']:
        messages.error(request, 'This import cannot be processed')
        return redirect('seller_tools:import_detail', pk=pk)

    # Trigger Celery task
    from seller_tools.tasks import process_bulk_import
    try:
        process_bulk_import.delay(bulk_import.id)
    except Exception:
        # Fallback to synchronous processing if Celery isn't running
        process_bulk_import(bulk_import.id)

    bulk_import.status = 'processing'
    bulk_import.started_at = timezone.now()
    bulk_import.save()

    messages.success(request, 'Import processing started. This may take a few minutes.')
    return redirect('seller_tools:import_detail', pk=pk)


@login_required
def download_import_template(request):
    """Download Excel template for bulk import with dropdown validation"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # --- Sheet 1: Listings (data entry) ---
    ws = wb.active
    ws.title = 'Listings'

    headers = [
        'title', 'description', 'category', 'condition', 'price', 'quantity',
        'grading_service', 'grade', 'cert_number', 'shipping_price',
        'listing_type', 'auction_duration_days', 'allow_offers',
        'image1_url', 'image2_url', 'image3_url', 'image4_url', 'image5_url',
        'video_url',
    ]
    col_widths = [35, 50, 20, 14, 12, 10, 16, 8, 14, 14, 14, 20, 14, 35, 35, 35, 35, 35, 35]

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2B3035', end_color='2B3035', fill_type='solid')
    thin_border = Border(
        bottom=Side(style='thin', color='DDDDDD'),
    )

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Example row
    example = [
        '1986 Fleer Michael Jordan #57 PSA 10',
        'Beautiful PSA 10 gem mint example of the iconic rookie card.',
        'trading-cards', 'mint', '500000.00', '1',
        'psa', '10', '12345678', '25.00', 'fixed', '', 'no',
        'https://example.com/front.jpg', 'back-photo.jpg', '', '', '',
    ]
    example_font = Font(color='888888', italic=True)
    for col_idx, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = example_font
        cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Reference ---
    ref = wb.create_sheet('Reference')
    ref_header_font = Font(bold=True, color='FFFFFF', size=11)
    ref_header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')

    # Column A: Category Name, Column B: Category Slug
    ref.cell(row=1, column=1, value='Category Name').font = ref_header_font
    ref.cell(row=1, column=1).fill = ref_header_fill
    ref.cell(row=1, column=2, value='Category Slug (use this)').font = ref_header_font
    ref.cell(row=1, column=2).fill = ref_header_fill
    ref.column_dimensions['A'].width = 30
    ref.column_dimensions['B'].width = 25

    categories = Category.objects.filter(is_active=True).order_by('name')
    cat_slugs = []
    for row_idx, cat in enumerate(categories, 2):
        ref.cell(row=row_idx, column=1, value=cat.name)
        ref.cell(row=row_idx, column=2, value=cat.slug)
        cat_slugs.append(cat.slug)
    cat_last_row = 1 + len(cat_slugs)

    # Column D: Conditions
    conditions = ['mint', 'near_mint', 'excellent', 'very_good', 'good', 'fair', 'poor']
    ref.cell(row=1, column=4, value='Condition').font = ref_header_font
    ref.cell(row=1, column=4).fill = ref_header_fill
    ref.column_dimensions['D'].width = 16
    for i, cond in enumerate(conditions, 2):
        ref.cell(row=i, column=4, value=cond)

    # Column F: Grading Services
    grading_services = ['psa', 'bgs', 'cgc', 'sgc']
    ref.cell(row=1, column=6, value='Grading Service').font = ref_header_font
    ref.cell(row=1, column=6).fill = ref_header_fill
    ref.column_dimensions['F'].width = 18
    for i, gs in enumerate(grading_services, 2):
        ref.cell(row=i, column=6, value=gs)

    # Column H: Listing Types
    ref.cell(row=1, column=8, value='Listing Type').font = ref_header_font
    ref.cell(row=1, column=8).fill = ref_header_fill
    ref.column_dimensions['H'].width = 14
    ref.cell(row=2, column=8, value='fixed')
    ref.cell(row=3, column=8, value='auction')

    # Column J: Allow Offers
    ref.cell(row=1, column=10, value='Allow Offers').font = ref_header_font
    ref.cell(row=1, column=10).fill = ref_header_fill
    ref.column_dimensions['J'].width = 14
    ref.cell(row=2, column=10, value='yes')
    ref.cell(row=3, column=10, value='no')

    # Column L: Image URL help
    ref.cell(row=1, column=12, value='Image Columns Help').font = ref_header_font
    ref.cell(row=1, column=12).fill = ref_header_fill
    ref.column_dimensions['L'].width = 45
    ref.cell(row=2, column=12, value='image1_url through image5_url accept:')
    ref.cell(row=3, column=12, value='  - Web URLs: https://example.com/photo.jpg')
    ref.cell(row=4, column=12, value='  - Filenames: my-photo.jpg (upload files with your Excel)')

    # --- Data Validation on Listings sheet ---
    # Category dropdown (column C, rows 2-500)
    if cat_slugs:
        dv_cat = DataValidation(
            type='list',
            formula1=f"Reference!$B$2:$B${cat_last_row}",
            allow_blank=True,
        )
        dv_cat.error = 'Please select a valid category from the Reference sheet'
        dv_cat.errorTitle = 'Invalid Category'
        dv_cat.prompt = 'Pick a category slug from the dropdown'
        dv_cat.promptTitle = 'Category'
        ws.add_data_validation(dv_cat)
        dv_cat.add(f'C2:C500')

    # Condition dropdown (column D)
    dv_cond = DataValidation(
        type='list',
        formula1=f"Reference!$D$2:$D${1 + len(conditions)}",
        allow_blank=True,
    )
    dv_cond.error = 'Please select a valid condition'
    dv_cond.errorTitle = 'Invalid Condition'
    ws.add_data_validation(dv_cond)
    dv_cond.add('D2:D500')

    # Grading service dropdown (column G)
    dv_grade = DataValidation(
        type='list',
        formula1=f"Reference!$F$2:$F${1 + len(grading_services)}",
        allow_blank=True,
    )
    dv_grade.error = 'Please select a valid grading service'
    dv_grade.errorTitle = 'Invalid Grading Service'
    ws.add_data_validation(dv_grade)
    dv_grade.add('G2:G500')

    # Listing type dropdown (column K)
    dv_type = DataValidation(
        type='list',
        formula1="Reference!$H$2:$H$3",
        allow_blank=True,
    )
    ws.add_data_validation(dv_type)
    dv_type.add('K2:K500')

    # Allow offers dropdown (column M)
    dv_offers = DataValidation(
        type='list',
        formula1="Reference!$J$2:$J$3",
        allow_blank=True,
    )
    ws.add_data_validation(dv_offers)
    dv_offers.add('M2:M500')

    # Write to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="listing_import_template.xlsx"'
    return response


@login_required
def inventory_list(request):
    """List inventory items"""
    items = InventoryItem.objects.filter(user=request.user)

    # Filters
    is_listed = request.GET.get('listed')
    if is_listed == 'yes':
        items = items.filter(is_listed=True)
    elif is_listed == 'no':
        items = items.filter(is_listed=False)

    category = request.GET.get('category')
    if category:
        items = items.filter(category__slug=category)

    return render(request, 'seller_tools/inventory_list.html', {
        'items': items,
        'categories': Category.objects.filter(parent__isnull=True),
    })


@login_required
def inventory_add(request):
    """Add inventory item"""
    if request.method == 'POST':
        item = InventoryItem.objects.create(
            user=request.user,
            title=request.POST.get('title'),
            category_id=request.POST.get('category') or None,
            condition=request.POST.get('condition', ''),
            grading_company=request.POST.get('grading_company', ''),
            grade=request.POST.get('grade') or None,
            cert_number=request.POST.get('cert_number', ''),
            purchase_price=request.POST.get('purchase_price') or None,
            purchase_date=request.POST.get('purchase_date') or None,
            purchase_source=request.POST.get('purchase_source', ''),
            target_price=request.POST.get('target_price') or None,
            minimum_price=request.POST.get('minimum_price') or None,
            notes=request.POST.get('notes', ''),
        )

        # Handle images
        for i in range(1, 4):
            img_field = f'image{i}'
            if img_field in request.FILES:
                setattr(item, img_field, request.FILES[img_field])
        item.save()

        messages.success(request, 'Inventory item added')
        return redirect('seller_tools:inventory_list')

    categories = Category.objects.filter(parent__isnull=True)

    return render(request, 'seller_tools/inventory_add.html', {
        'categories': categories,
    })


@login_required
def inventory_detail(request, pk):
    """View inventory item"""
    item = get_object_or_404(InventoryItem, pk=pk, user=request.user)

    return render(request, 'seller_tools/inventory_detail.html', {
        'item': item,
    })


@login_required
def inventory_edit(request, pk):
    """Edit inventory item"""
    item = get_object_or_404(InventoryItem, pk=pk, user=request.user)

    if request.method == 'POST':
        item.title = request.POST.get('title')
        item.category_id = request.POST.get('category') or None
        item.condition = request.POST.get('condition', '')
        item.grading_company = request.POST.get('grading_company', '')
        item.grade = request.POST.get('grade') or None
        item.cert_number = request.POST.get('cert_number', '')
        item.purchase_price = request.POST.get('purchase_price') or None
        item.purchase_date = request.POST.get('purchase_date') or None
        item.purchase_source = request.POST.get('purchase_source', '')
        item.target_price = request.POST.get('target_price') or None
        item.minimum_price = request.POST.get('minimum_price') or None
        item.notes = request.POST.get('notes', '')
        item.save()

        messages.success(request, 'Inventory item updated')
        return redirect('seller_tools:inventory_detail', pk=pk)

    categories = Category.objects.filter(parent__isnull=True)

    return render(request, 'seller_tools/inventory_edit.html', {
        'item': item,
        'categories': categories,
    })


@login_required
def inventory_create_listing(request, pk):
    """Create listing from inventory item"""
    item = get_object_or_404(InventoryItem, pk=pk, user=request.user)

    # Pre-populate session data for listing create
    request.session['listing_prefill'] = {
        'title': item.title,
        'category': item.category_id,
        'condition': item.condition,
        'grading_service': item.grading_company,
        'grade': str(item.grade) if item.grade else '',
        'cert_number': item.cert_number,
        'price': str(item.target_price) if item.target_price else '',
        'price_guide_item': item.price_guide_item_id,
    }

    messages.info(request, 'Complete your listing with additional details.')
    return redirect('marketplace:listing_create')


@login_required
def seller_analytics(request):
    """Seller analytics dashboard"""
    # Date range
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)

    # Sales over time
    sales_by_day = Order.objects.filter(
        seller=request.user,
        status__in=['paid', 'shipped', 'delivered', 'completed'],
        created__gte=start_date
    ).annotate(
        day=TruncDay('created')
    ).values('day').annotate(
        total=Sum('item_price'),
        count=Count('id')
    ).order_by('day')

    # Top selling items
    top_items = Order.objects.filter(
        seller=request.user,
        status__in=['paid', 'shipped', 'delivered', 'completed'],
        created__gte=start_date
    ).values(
        'listing__title', 'listing__category__name'
    ).annotate(
        total=Sum('item_price'),
        count=Count('id')
    ).order_by('-total')[:10]

    # Total stats
    totals = Order.objects.filter(
        seller=request.user,
        status__in=['paid', 'shipped', 'delivered', 'completed'],
        created__gte=start_date
    ).aggregate(
        revenue=Sum('item_price'),
        orders=Count('id'),
        fees=Sum('platform_fee')
    )

    return render(request, 'seller_tools/analytics.html', {
        'sales_by_day': list(sales_by_day),
        'top_items': top_items,
        'totals': totals,
        'days': days,
    })


@login_required
def sales_report(request):
    """Detailed sales report"""
    # Date range
    start = request.GET.get('start')
    end = request.GET.get('end')

    orders = Order.objects.filter(
        seller=request.user,
        status__in=['paid', 'shipped', 'delivered', 'completed']
    ).order_by('-created')

    if start:
        orders = orders.filter(created__date__gte=start)
    if end:
        orders = orders.filter(created__date__lte=end)

    totals = orders.aggregate(
        revenue=Sum('item_price'),
        fees=Sum('platform_fee'),
        net=Sum('seller_payout')
    )

    return render(request, 'seller_tools/sales_report.html', {
        'orders': orders[:100],
        'totals': totals,
    })


@login_required
def export_analytics(request):
    """Export sales data as CSV"""
    orders = Order.objects.filter(
        seller=request.user,
        status__in=['paid', 'shipped', 'delivered', 'completed']
    ).order_by('-created')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sales_export.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Order ID', 'Date', 'Item', 'Buyer', 'Item Price', 'Shipping',
        'Total', 'Platform Fee', 'Your Payout', 'Status'
    ])

    for order in orders:
        writer.writerow([
            order.id,
            order.created.strftime('%Y-%m-%d %H:%M'),
            order.listing.title if order.listing else 'Deleted',
            order.buyer.username,
            order.item_price,
            order.shipping_price,
            order.amount,
            order.platform_fee,
            order.seller_payout,
            order.status,
        ])

    return response


@login_required
def payout_settings(request):
    """Seller payout settings and history"""
    from marketplace.services.connect_service import ConnectService

    profile = request.user.profile

    if not profile.stripe_account_id:
        messages.info(request, 'Please complete your seller account setup first.')
        return redirect('marketplace:seller_setup')

    # Get account details
    try:
        account = ConnectService.retrieve_account(profile.stripe_account_id)
        balance = ConnectService.get_balance(profile.stripe_account_id)
        transfers = ConnectService.list_transfers(profile.stripe_account_id, limit=20)
    except Exception as e:
        messages.error(request, f'Unable to load payout information: {e}')
        account = None
        balance = {'available': [], 'pending': []}
        transfers = []

    return render(request, 'seller_tools/payout_settings.html', {
        'account': account,
        'balance': balance,
        'transfers': transfers.data if hasattr(transfers, 'data') else [],
    })


# --- Photo Capture Flow ---

def _get_import_listings(bulk_import):
    """Get all listings created by a bulk import."""
    row_ids = bulk_import.rows.filter(
        status='success', listing__isnull=False
    ).values_list('listing_id', flat=True)
    return Listing.objects.filter(id__in=row_ids).order_by('id')


@login_required
def import_photos(request, pk):
    """Overview grid of all listings from an import needing photos."""
    bulk_import = get_object_or_404(BulkImport, pk=pk, user=request.user)
    listings = _get_import_listings(bulk_import)

    listings_data = []
    with_photos = 0
    for listing in listings:
        photo_count = sum(1 for i in range(1, 6) if getattr(listing, f'image{i}'))
        if photo_count > 0:
            with_photos += 1
        listings_data.append({
            'listing': listing,
            'photo_count': photo_count,
            'thumbnail': getattr(listing, 'image1') if getattr(listing, 'image1') else None,
        })

    total = len(listings_data)
    progress = int((with_photos / total) * 100) if total else 0

    return render(request, 'seller_tools/import_photos.html', {
        'bulk_import': bulk_import,
        'listings_data': listings_data,
        'total': total,
        'with_photos': with_photos,
        'progress': progress,
    })


@login_required
def import_photo_capture(request, pk, listing_id):
    """Mobile-first photo capture page for a single listing."""
    bulk_import = get_object_or_404(BulkImport, pk=pk, user=request.user)
    listing = get_object_or_404(Listing, pk=listing_id, seller=request.user)

    # Verify listing belongs to this import
    if not bulk_import.rows.filter(listing=listing).exists():
        messages.error(request, 'Listing does not belong to this import.')
        return redirect('seller_tools:import_photos', pk=pk)

    # Build photo slots
    slots = []
    for i in range(1, 6):
        field = f'image{i}'
        image = getattr(listing, field)
        slots.append({
            'position': i,
            'label': 'Main' if i == 1 else f'Photo {i}',
            'image': image if image else None,
        })

    # Navigation: prev/next listings in this import
    all_listings = list(_get_import_listings(bulk_import).values_list('id', flat=True))
    current_idx = all_listings.index(listing.id) if listing.id in all_listings else 0
    prev_id = all_listings[current_idx - 1] if current_idx > 0 else None
    next_id = all_listings[current_idx + 1] if current_idx < len(all_listings) - 1 else None

    # Find next listing needing photos (batch query)
    no_photo_ids = set(
        Listing.objects.filter(id__in=all_listings, image1='')
        .values_list('id', flat=True)
    ) | set(
        Listing.objects.filter(id__in=all_listings, image1__isnull=True)
        .values_list('id', flat=True)
    )
    next_needing = None
    # Look after current first, then wrap around
    for lid in all_listings[current_idx + 1:] + all_listings[:current_idx]:
        if lid in no_photo_ids:
            next_needing = lid
            break

    return render(request, 'seller_tools/import_photo_capture.html', {
        'bulk_import': bulk_import,
        'listing': listing,
        'slots': slots,
        'prev_id': prev_id,
        'next_id': next_id,
        'next_needing': next_needing,
        'current_num': current_idx + 1,
        'total_num': len(all_listings),
    })


@login_required
@require_POST
def import_photo_upload(request, pk, listing_id):
    """HTMX endpoint: upload a photo to a listing slot."""
    bulk_import = get_object_or_404(BulkImport, pk=pk, user=request.user)
    listing = get_object_or_404(Listing, pk=listing_id, seller=request.user)

    if not bulk_import.rows.filter(listing=listing).exists():
        return HttpResponse('Unauthorized', status=403)

    if 'photo' not in request.FILES:
        return HttpResponse('No file', status=400)

    photo = request.FILES['photo']

    # Find the target position (explicit via query param/POST or first empty)
    position = request.GET.get('position') or request.POST.get('position')
    if position:
        position = int(position)
    else:
        position = None
        for i in range(1, 6):
            if not getattr(listing, f'image{i}'):
                position = i
                break
        if position is None:
            return HttpResponse('All slots full', status=400)

    field = f'image{position}'
    setattr(listing, field, photo)
    listing.save(update_fields=[field])

    image = getattr(listing, field)
    return render(request, 'seller_tools/partials/photo_slot.html', {
        'slot': {
            'position': position,
            'label': 'Main' if position == 1 else f'Photo {position}',
            'image': image,
        },
        'bulk_import': bulk_import,
        'listing': listing,
    })


@login_required
@require_POST
def import_photo_delete(request, pk, listing_id, position):
    """HTMX endpoint: delete a photo from a listing slot."""
    bulk_import = get_object_or_404(BulkImport, pk=pk, user=request.user)
    listing = get_object_or_404(Listing, pk=listing_id, seller=request.user)

    if not bulk_import.rows.filter(listing=listing).exists():
        return HttpResponse('Unauthorized', status=403)

    if position < 1 or position > 5:
        return HttpResponse('Invalid position', status=400)

    field = f'image{position}'
    image_field = getattr(listing, field)
    if image_field:
        image_field.delete(save=False)
    setattr(listing, field, '')
    listing.save(update_fields=[field])

    return render(request, 'seller_tools/partials/photo_slot.html', {
        'slot': {
            'position': position,
            'label': 'Main' if position == 1 else f'Photo {position}',
            'image': None,
        },
        'bulk_import': bulk_import,
        'listing': listing,
    })


@login_required
def ship_from_address(request):
    """Manage seller's ship-from (return) address."""
    from shipping.models import Address

    profile = request.user.profile
    address = profile.default_ship_from

    if request.method == 'POST':
        addr_data = {
            'name': request.POST.get('name', '').strip(),
            'company': request.POST.get('company', '').strip(),
            'street1': request.POST.get('street1', '').strip(),
            'street2': request.POST.get('street2', '').strip(),
            'city': request.POST.get('city', '').strip(),
            'state': request.POST.get('state', '').strip(),
            'zip_code': request.POST.get('zip_code', '').strip(),
            'country': request.POST.get('country', 'US').strip(),
            'phone': request.POST.get('phone', '').strip(),
        }

        if not all([addr_data['name'], addr_data['street1'], addr_data['city'],
                     addr_data['state'], addr_data['zip_code']]):
            messages.error(request, 'Please fill in all required fields.')
            return render(request, 'seller_tools/ship_from_address.html', {
                'address': addr_data,
            })

        # Verify via EasyPost if configured
        verified = False
        easypost_id = ''
        if settings.EASYPOST_API_KEY or settings.USPS_CLIENT_ID:
            from marketplace.services.shipping_factory import get_shipping_service
            ShippingService = get_shipping_service()
            result = ShippingService.verify_address({
                'name': addr_data['name'],
                'company': addr_data.get('company'),
                'street1': addr_data['street1'],
                'street2': addr_data.get('street2'),
                'city': addr_data['city'],
                'state': addr_data['state'],
                'zip': addr_data['zip_code'],
                'country': addr_data['country'],
            })
            if result['verified']:
                verified = True
                easypost_id = result['easypost_id']
                # Use corrected address
                corrected = result['address']
                addr_data['street1'] = corrected['street1']
                addr_data['street2'] = corrected.get('street2', '')
                addr_data['city'] = corrected['city']
                addr_data['state'] = corrected['state']
                addr_data['zip_code'] = corrected['zip']

        if address:
            for key, val in addr_data.items():
                setattr(address, key, val)
            address.is_verified = verified
            address.easypost_id = easypost_id
            address.save()
        else:
            address = Address.objects.create(
                user=request.user,
                is_verified=verified,
                easypost_id=easypost_id,
                **addr_data,
            )
            profile.default_ship_from = address
            profile.save(update_fields=['default_ship_from'])

        messages.success(request, 'Ship-from address saved!' +
                         (' (verified)' if verified else ''))
        return redirect('seller_tools:ship_from_address')

    context = {
        'address': address,
    }
    return render(request, 'seller_tools/ship_from_address.html', context)
